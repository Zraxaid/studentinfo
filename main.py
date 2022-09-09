import openpyxl
import os
import datetime
from openpyxl import styles
from copy import copy

# get the path for the root directory
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))

# concatenate the path of the root dir and the calendar excel file
excel_file = os.path.join(ROOT_DIR, "Student Info.xlsx")

workbook = openpyxl.load_workbook(excel_file)

curr_month = datetime.datetime.now().month
month_ref = {1: "February",
             2: "March",
             3: "April",
             4: "May",
             5: "June",
             6: "July",
             7: "August",
             8: "September",
             9: "October",
             10: "November",
             11: "December",
             12: "January",
             0: "January"}
curr_sheet = workbook[month_ref[curr_month - 1]]


class Teacher:
    def __init__(self, name, color):
        self.name = name
        self.students = []
        self.color = color

    def __str__(self):
        return f"[Name: {str(self.name)}  Color: {str(self.color)}  Students: {str(self.students)}]"

    def __repr__(self):
        return f"[Name: {str(self.name)}  Color: {str(self.color)}  Students: {str(self.students)}]"

    def add_student(self, student):
        self.students.append(student)


class Student:
    def __init__(self, name, subject, parent, credit, time_of_week, row_num):
        self.name = name
        self.subject = subject
        self.classes = []
        self.class_num = 0
        self.parent = parent
        self.credit = credit
        self.time_of_week = time_of_week
        self.row_num = row_num

    def add_class(self, date):
        # credit - class num
        self.class_num += 1
        self.classes.append(date)

    def __str__(self):
        return f"Name: {self.name}\n      Subject: {self.subject}\n      Parent Name: {self.parent}\n      " \
               f"Classes({self.class_num}): {self.classes}\n      Leftover Credit: {self.credit - self.class_num} "

    def __repr__(self):
        return f"[Name: {self.name} Subject: {self.subject} Parent Name: {self.parent} " \
               f"Classes({self.class_num}): {self.classes}]"


def print_teachers(teacher_list, choice):
    # teacher_list is an array of teacher object, choice is a string (either the teachers name or "all")
    if choice == "all":
        for i in teacher_list:
            print(i.name)
            for j in i.students:
                print(f"   {j}")
    else:
        for i in teacher_list:
            if i.name == choice:
                print(i.name)
                for j in i.students:
                    print(f"   {j}")


def init_teachers(working_sheet):
    column_num = 0
    dict1 = []
    dict2 = {}
    # looks for teacher and color cells
    for i in range(1, working_sheet.max_column + 1):
        if working_sheet.cell(row=1, column=i).value == "Teacher" and working_sheet.cell(row=1,
                                                                                         column=i + 1).value == "Color":
            column_num = i
            break
    # creates teachers and assigns them colors
    for i in range(2, working_sheet.max_row + 1):
        if working_sheet.cell(row=i, column=column_num).value is None:
            break
        # get teacher color on excel
        result = working_sheet.cell(row=i, column=column_num + 1).fill.start_color.index
        dict1.append(Teacher(working_sheet.cell(row=i, column=column_num).value, result))
        dict2[result] = dict1[-1]
    return dict1, dict2


def assign_students(working_sheet, teacher_dict):
    column_num = 0
    row_num = 0
    # go through columns to see what columns to loop through
    for i in range(1, working_sheet.max_column + 1):
        if working_sheet.cell(row=1, column=i).value is None:
            column_num = i
            break
    # go through rows to see what columns to loop through (at some point the rows are empty)
    for i in range(2, working_sheet.max_row + 1):
        if working_sheet.cell(row=i, column=1).value is None:
            row_num = i
            break
    # loop through rows to get students and assign them to teachers
    for i in range(2, row_num):
        current_color = working_sheet.cell(row=i, column=1).fill.start_color.index
        if current_color in teacher_dict.keys():
            # create student with their name, parent, subject, and credit [look to student class __init__]
            new_student = Student(working_sheet.cell(row=i, column=3).value, working_sheet.cell(row=i, column=1).value,
                                  working_sheet.cell(row=i, column=2).value, working_sheet.cell(row=i, column=5).value,
                                  working_sheet.cell(row=i, column=4).value, i)
            # record current classes
            for j in range(7, column_num):
                if working_sheet.cell(row=i, column=j).value is not None:
                    try:
                        date = f"{working_sheet.cell(row=i, column=j).value.month}/{working_sheet.cell(row=i, column=j).value.day}/{working_sheet.cell(row=i, column=j).value.year}"
                        new_student.add_class(date)
                    except:
                        date= f"{working_sheet.cell(row=i, column=j).value.split('/')[0]}/{working_sheet.cell(row=i, column=j).value.split('/')[1]}/{working_sheet.cell(row=i, column=j).value.split('/')[2]}"
                        new_student.add_class(date)
                else:
                    break
            # assign student to their teacher
            teacher_dict[current_color].add_student(new_student)
        else:
            print(f"{working_sheet.cell(row=i, column=3).value} has no color")


def new_month(entire_sheet):
    print(month_ref[curr_month])
    template_sheet = entire_sheet[month_ref[curr_month-1]]
    entire_sheet.create_sheet(month_ref[curr_month])
    entire_sheet.save(filename="Student Info.xlsx")
    current_working_sheet = entire_sheet[month_ref[curr_month]]
    for i in range(1, template_sheet.max_column + 1):
        for j in range(1, template_sheet.max_row + 1):
            current_working_sheet.cell(row=j, column=i).value = template_sheet.cell(row=j, column=i).value
            if template_sheet.cell(row=j, column=i).has_style:
                current_working_sheet.cell(row=j, column=i).font = copy(template_sheet.cell(row=j, column=i).font)
                current_working_sheet.cell(row=j, column=i).border = copy(template_sheet.cell(row=j, column=i).border)
                current_working_sheet.cell(row=j, column=i).fill = copy(template_sheet.cell(row=j, column=i).fill)
                current_working_sheet.cell(row=j, column=i).number_format = copy(template_sheet.cell(row=j, column=i).number_format)
                current_working_sheet.cell(row=j, column=i).protection = copy(template_sheet.cell(row=j, column=i).protection)
                current_working_sheet.cell(row=j, column=i).alignment = copy(template_sheet.cell(row=j, column=i).alignment)
    for i in "A B C D E F G H I J K L M N O P Q R S T U V W X Y Z".split(" "):
        current_working_sheet.column_dimensions[i].width = template_sheet.column_dimensions[i].width
    for i in range(1, template_sheet.max_row + 1):
        current_working_sheet.row_dimensions[i].height = template_sheet.row_dimensions[i].height
    for j in range(2, template_sheet.max_row + 1):
        current_working_sheet.cell(row=j, column=5).value = template_sheet.cell(row=j, column=6).value
        current_working_sheet.cell(row=j, column=7).value = None
        current_working_sheet.cell(row=j, column=8).value = None
        current_working_sheet.cell(row=j, column=9).value = None
        current_working_sheet.cell(row=j, column=10).value = None
        current_working_sheet.cell(row=j, column=11).value = None
        current_working_sheet.cell(row=j, column=12).value = None
        current_working_sheet.cell(row=j, column=13).value = None
        current_working_sheet.cell(row=j, column=14).value = None
        current_working_sheet.cell(row=j, column=15).value = None
    entire_sheet.save(filename="Student Info.xlsx")


def finished_class(teacher_dictionary, input_teacher):
    # let user input the student and subject and mark today's class as done and -1 credit
    date = f"{datetime.datetime.now().month}/{datetime.datetime.now().day}/{datetime.datetime.now().year}"
    print(f"Today's Date: {date}")
    print_teachers(teacher_dictionary, input_teacher.name)
    confirmed_student = None
    while not confirmed_student:
        selected_student = input("Please enter the name of the student: ").lower()
        selected_subject = input("Please enter the student's subject: ").lower()
        for i in input_teacher.students:
            # confirms the student and the subject
            if i.name.lower() == selected_student and i.subject.lower() == selected_subject:
                print(f"You selected: {i.name} ({i.subject})")
                confirm_student = input("Confirm (y/n): ")
                if confirm_student == "y":
                    confirmed_student = i
    confirmed_student.add_class(date)
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for i in range(7, 16):
        if workbook[workbook.sheetnames[-2]].cell(row=confirmed_student.row_num, column=i).value is None:
            curr_sheet[alphabet[i-1]+str(confirmed_student.row_num)] = date
            break
        workbook[workbook.sheetnames[-2]].cell(row=confirmed_student.row_num, column=6).value -= 1
    workbook.save(filename="Student Info.xlsx")


def add_classes(teacher_dictionary, input_teacher):
    confirmed_student = None
    print_teachers(teacher_dictionary, input_teacher.name)
    while not confirmed_student:
        selected_student = input("Please enter the name of the student: ").lower()
        selected_subject = input("Please enter the student's subject: ").lower()
        credit = int(input("Please enter the number of classes: "))
        for i in input_teacher.students:
            # confirms the student and the subject
            if i.name.lower() == selected_student and i.subject.lower() == selected_subject:
                print(f"You want to add {credit} classes to {i.name} ({i.subject})")
                confirm_student = input("Confirm (y/n): ")
                if confirm_student == "y":
                    confirmed_student = i
    workbook[workbook.sheetnames[-2]].cell(row=confirmed_student.row_num, column=5).value += credit
    workbook[workbook.sheetnames[-2]].cell(row=confirmed_student.row_num, column=6).value += credit
    workbook.save(filename="Student Info.xlsx")


def operation(teacher_dict):
    # main loop asking for inputs
    teacher = None
    while not teacher:
        selected_teacher = input("Please enter the name of the teacher: ")
        selected_teacher = selected_teacher[0].upper() + selected_teacher[1:].lower()
        print("You selected:")
        print("--------------------------------------------")
        print_teachers(teacher_dict, selected_teacher)
        print("--------------------------------------------")
        confirm = input("Confirm (y/n): ")
        if confirm == "y":
            for i in teacher_dict:
                if i.name == selected_teacher:
                    teacher = i
    while True:
        print("--------------------------------------------")
        print(f"You selected: {teacher.name}")
        print('''What would you like to do?\n
        1. Finished Class\n
        2. Add Credit\n
        3. Report Status\n
        4. New Month\n
        quit: exit program''')
        print("--------------------------------------------")
        choice = int(input("Choice: ").strip(" "))
        if choice == "quit":
            break
        else:
            print(f"You chose: {choice}")
            print("--------------------------------------------")
            choice_dictionary = {1: "finished_class(teacher_dict, teacher)",
                                 2: "add_classes(teacher_dict, teacher)",
                                 4: "new_month(workbook)"}
            exec(choice_dictionary[choice])
            print("--------------------------------------------")
            print_teachers(teacher_dict, teacher.name)


# initialize everything
print("--------------------------------------------")
print("Initializing...")
Teacher_Dict, color_to_teacher = init_teachers(curr_sheet)
assign_students(curr_sheet, color_to_teacher)
print("Initialization complete!")
print(f"Current month: {month_ref[curr_month-1]}")
print("--------------------------------------------")
# print_teachers(Teacher_Dict, "all")

operation(Teacher_Dict)
